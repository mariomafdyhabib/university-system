import re
import openpyxl
import pandas as pd
from database import db, Courses, Instructors, CourseSections, Schedules

# ── Day-name short → full ────────────────────────────────────────────────────
DAY_MAP = {
    'Sun': 'Sunday', 'Mon': 'Monday', 'Tue': 'Tuesday',
    'Wed': 'Wednesday', 'Thu': 'Thursday', 'Fri': 'Friday', 'Sat': 'Saturday',
}


def _norm_time(t):
    """Normalise a time string to '8:00 AM' / '1:30 PM' format."""
    t = str(t).strip()
    m = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)?', t, re.IGNORECASE)
    if not m:
        return t
    h, mn, ampm = int(m.group(1)), m.group(2), (m.group(3) or '').upper()
    if not ampm:
        ampm = 'AM' if h >= 7 else 'PM'
    return f"{h}:{mn} {ampm}"


def _parse_time_range(time_str):
    """'08:00 AM - 09:30 AM' → ('8:00 AM', '9:30 AM')"""
    try:
        parts = re.split(r'\s*-\s*', str(time_str).strip())
        if len(parts) < 2:
            return time_str.strip(), time_str.strip()
        return _norm_time(parts[0]), _norm_time(parts[1])
    except Exception:
        raw = str(time_str).strip()
        return raw, raw


def _parse_section_field(section_field):
    """
    'ACCT 130 - SEC 81'  → ('ACCT 130', 'SEC 81')
    'MEDI 401- SEC 81'   → ('MEDI 401', 'SEC 81')
    """
    if not section_field:
        return '', ''
    parts = re.split(r'\s*-\s*', str(section_field).strip(), maxsplit=1)
    code = parts[0].strip()
    sec  = parts[1].strip() if len(parts) > 1 else 'SEC 81'
    return code, sec


def _parse_days(days_str):
    """'Sun,Tue,Thu' → ['Sunday', 'Tuesday', 'Thursday']"""
    if not days_str:
        return []
    return [DAY_MAP.get(d.strip(), d.strip()) for d in str(days_str).split(',')]


def _is_schedule_report_format(columns):
    """Detect our Schedule_Report Excel format by column names."""
    cols = [str(c).strip() for c in columns]
    return 'Course Section' in cols or 'Day Name' in cols


# ── Main entry point called by admin_routes.py ───────────────────────────────

def process_upload(file_path, mode='append'):
    """
    Parse a CSV or Excel file and populate Courses / CourseSections / Schedules.

    Supports two Excel layouts:
      1. Schedule Report format  – columns: Course, Course Section, Room, Day Name, Time
      2. Generic CRN format      – columns: Course, Section, Day, Time, Room, Instructor

    mode='append'  → add rows on top of existing data (default)
    mode='replace' → wipe schedule tables first then repopulate
    """
    try:
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
            return _process_generic(df, mode)
        else:
            # Use openpyxl to detect sheet name
            wb = openpyxl.load_workbook(file_path)
            # Prefer the "Schedule_Report_Improved" sheet if present
            target_sheet = None
            for name in wb.sheetnames:
                if 'schedule' in name.lower() or 'report' in name.lower():
                    target_sheet = name
                    break
            if target_sheet is None:
                target_sheet = wb.sheetnames[0]

            df = pd.read_excel(file_path, sheet_name=target_sheet)

        if _is_schedule_report_format(df.columns):
            return _process_schedule_report(df, mode)
        else:
            return _process_generic(df, mode)

    except Exception as e:
        db.session.rollback()
        raise e


# ── Layout 1: Schedule_Report (Course / Course Section / Room / Day Name / Time) ──

def _process_schedule_report(df, mode):
    if mode == 'replace':
        Schedules.query.delete()
        CourseSections.query.delete()
        Courses.query.delete()
        db.session.flush()

    course_cache  = {}   # code   → Courses obj
    section_cache = {}   # (code, sec_label) → CourseSections obj
    rows_added = 0

    for _, row in df.iterrows():
        course_name   = str(row.get('Course', '')).strip()
        section_field = str(row.get('Course Section', '')).strip()
        room          = str(row.get('Room', '')).strip()
        days_raw      = str(row.get('Day Name', '')).strip()
        time_raw      = str(row.get('Time', '')).strip()

        if not section_field or not course_name or section_field == 'nan':
            continue

        code, sec_label = _parse_section_field(section_field)
        if not code:
            continue

        dept = code.split()[0] if ' ' in code else code[:4]
        start_time, end_time = _parse_time_range(time_raw)
        days = _parse_days(days_raw)

        # ── Course ────────────────────────────────────────────────────────
        if code not in course_cache:
            c = Courses.query.filter_by(course_code=code).first()
            if not c:
                c = Courses(course_code=code, course_name=course_name,
                            credits=3, department=dept)
                db.session.add(c)
                db.session.flush()
            course_cache[code] = c
        course = course_cache[code]

        # ── Section ───────────────────────────────────────────────────────
        sec_key = (code, sec_label)
        if sec_key not in section_cache:
            sec = CourseSections.query.filter_by(
                course_id=course.course_id, section_name=sec_label
            ).first()
            if not sec:
                sec = CourseSections(
                    course_id=course.course_id,
                    instructor_id=None,
                    section_name=sec_label,
                    semester='Spring 2026'
                )
                db.session.add(sec)
                db.session.flush()
            section_cache[sec_key] = sec
        sec = section_cache[sec_key]

        # ── Schedule rows (one per day) ───────────────────────────────────
        for day in days:
            exists = Schedules.query.filter_by(
                section_id=sec.section_id, day_of_week=day
            ).first()
            if not exists:
                db.session.add(Schedules(
                    section_id=sec.section_id,
                    day_of_week=day,
                    start_time=start_time,
                    end_time=end_time,
                    classroom=room
                ))
            rows_added += 1

    db.session.commit()
    return rows_added


# ── Layout 2: Generic CRN format (Course / Section / Day / Time / Room / Instructor) ──

def _process_generic(df, mode):
    if mode == 'replace':
        Schedules.query.delete()
        CourseSections.query.delete()
        Courses.query.delete()
        db.session.flush()

    course_cache  = {}
    inst_cache    = {}
    section_cache = {}
    rows_added = 0

    for _, row in df.iterrows():
        course_name  = str(row.get('Course', '')).strip()
        section_name = str(row.get('Section', '')).strip()
        day          = str(row.get('Day', '')).strip()
        time_raw     = str(row.get('Time', '')).strip()
        room         = str(row.get('Room', '')).strip()
        inst_name    = str(row.get('Instructor', '')).strip()

        if not course_name or not day or course_name == 'nan':
            continue

        start_time, end_time = _parse_time_range(time_raw)

        # Course
        if course_name not in course_cache:
            c = Courses.query.filter_by(course_name=course_name).first()
            if not c:
                words = course_name.split()
                code  = (words[0][:4] + (words[-1][:3] if len(words) > 1 else '')).upper()
                base, n = code, 1
                while Courses.query.filter_by(course_code=code).first():
                    code = f"{base}{n}"; n += 1
                c = Courses(course_name=course_name, course_code=code, credits=3)
                db.session.add(c)
                db.session.flush()
            course_cache[course_name] = c
        course = course_cache[course_name]

        # Instructor
        inst_id = None
        if inst_name and inst_name != 'nan':
            if inst_name not in inst_cache:
                inst = Instructors.query.filter_by(name=inst_name).first()
                if not inst:
                    inst = Instructors(name=inst_name)
                    db.session.add(inst)
                    db.session.flush()
                inst_cache[inst_name] = inst
            inst_id = inst_cache[inst_name].instructor_id

        # Section
        sec_key = (course.course_id, section_name)
        if sec_key not in section_cache:
            sec = CourseSections.query.filter_by(
                course_id=course.course_id, section_name=section_name
            ).first()
            if not sec:
                sec = CourseSections(
                    course_id=course.course_id,
                    instructor_id=inst_id,
                    section_name=section_name,
                    semester='TBD'
                )
                db.session.add(sec)
                db.session.flush()
            section_cache[sec_key] = sec
        sec = section_cache[sec_key]

        # Schedule
        exists = Schedules.query.filter_by(
            section_id=sec.section_id, day_of_week=day
        ).first()
        if not exists:
            db.session.add(Schedules(
                section_id=sec.section_id,
                day_of_week=day,
                start_time=start_time,
                end_time=end_time,
                classroom=room
            ))
        rows_added += 1

    db.session.commit()
    return rows_added
